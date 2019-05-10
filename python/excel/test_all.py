from openpyxl import Workbook,worksheet,styles
from openpyxl.styles import PatternFill,Border,Side,Alignment,Protection,Font
import platform 

wb = Workbook()
ws = wb.active
ws.title='name'
cell_font = Font(size=22,bold=True,color='FF00FF')

ws['a1'].value = 'aa'
for i in range(10):
    ws['a%d' %(i+1)].value = i+1
    if i % 2 == 0:
        ws['a%d' % (i + 1)].font = cell_font

print(platform.python_version())

wb.save(r'/Users/miaozw/Documents/TEMP/python/test.xlsx')
