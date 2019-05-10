#!/usr/local/bin/python3

from openpyxl import Workbook,worksheet,styles
from openpyxl.styles import PatternFill,Border,Side,Alignment,Protection,Font
import platform 
import openpyxl
import os

wb = Workbook()
ws = wb.active
ws.title='name'
cell_font = Font(size=22,bold=True,color='FF00FF')
cell_font.size=32

ws['a1'].value = 'aa'
for i in range(10):
    ws['a%d' %(i+1)].value = i+1
    if i % 2 == 0:
        ws['a%d' % (i + 1)].font = cell_font

print(openpyxl.__version__)

wb.save(r'/Users/miaozw/Documents/TEMP/python/test.xlsx')

# os.system("open /Users/miaozw/Documents/TEMP/python/test.xlsx")
